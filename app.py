"""
Sistema de Consulta Tributária - IBS/CBS
VERSÃO APRIMORADA com melhorias didáticas, busca inteligente e filtros avançados

Melhorias implementadas:
- Busca com sinônimos e palavras-chave
- Busca por código LC116 e NBS
- Correspondência parcial e normalização de acentos
- Autocompletar inteligente
- Filtros didáticos com linguagem clara
- Destaque de termos buscados
- Cores e ícones para categorias
- Ordenação de resultados
"""

import streamlit as st
import pandas as pd
from pathlib import Path
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Importar serviços
from services.data_service import DataService
from services.search_service import SearchServiceEnhanced, GRUPOS_LC116

# =============================================================================
# CONFIGURAÇÕES
# =============================================================================

# Caminhos
DATA_FILE = Path(__file__).parent / "data" / "anexoVIII_correlacao_categorizado.json"

# Configurações de busca
SEARCH_CONFIG = {
    "fuzzy_threshold": 65,
    "min_search_length": 2,
    "max_autocomplete": 8,
}


# =============================================================================
# CONFIGURAÇÃO DA PÁGINA
# =============================================================================

def configure_page():
    """Configura a página Streamlit com tema premium."""
    st.set_page_config(
        page_title="Consulta Tributária IBS/CBS - Neto Contabilidade",
        page_icon="📊",
        layout="wide",
        initial_sidebar_state="expanded"
    )

    # CSS Premium com melhorias visuais
    st.markdown("""
    <style>
        /* ============================================
           TEMA BASE - DARK PREMIUM
           ============================================ */
        .stApp {
            background: linear-gradient(135deg, #1a2332 0%, #2d3e50 100%);
            color: #e6edf3;
        }

        /* ============================================
           HEADER PREMIUM
           ============================================ */
        .premium-header {
            background: linear-gradient(135deg, rgba(201, 169, 97, 0.15) 0%, rgba(230, 213, 168, 0.1) 100%);
            border: 1px solid rgba(201, 169, 97, 0.3);
            border-radius: 16px;
            padding: 25px 35px;
            margin-bottom: 30px;
            display: flex;
            align-items: center;
            gap: 20px;
        }

        .logo-circle {
            width: 60px;
            height: 60px;
            background: linear-gradient(135deg, #c9a961 0%, #e6d5a8 100%);
            border-radius: 50%;
            display: flex;
            align-items: center;
            justify-content: center;
            font-weight: 700;
            font-size: 22px;
            color: #1a2332;
            box-shadow: 0 4px 15px rgba(201, 169, 97, 0.4);
        }

        .header-title {
            font-size: 28px;
            font-weight: 700;
            color: #fff;
            margin-bottom: 5px;
        }

        .header-subtitle {
            color: #c9a961;
            font-size: 14px;
            letter-spacing: 1px;
        }

        /* ============================================
           HERO DE BUSCA
           ============================================ */
        .search-hero {
            text-align: center;
            padding: 40px 20px 30px;
        }

        .search-title {
            font-size: 32px;
            font-weight: 700;
            color: #fff;
            margin-bottom: 10px;
        }

        .search-subtitle {
            color: #b0b8c1;
            margin-bottom: 25px;
            font-size: 16px;
        }

        /* ============================================
           CAMPO DE BUSCA ESTILIZADO
           ============================================ */
        .stTextInput input,
        .stTextInput > div > div > input,
        [data-testid="stTextInput"] input {
            background: rgba(255,255,255,0.1) !important;
            background-color: rgba(255,255,255,0.1) !important;
            border: 2px solid rgba(201, 169, 97, 0.3) !important;
            border-radius: 50px !important;
            padding: 15px 25px !important;
            color: #ffffff !important;
            font-size: 16px !important;
            caret-color: #c9a961 !important;
        }

        .stTextInput input:focus,
        .stTextInput > div > div > input:focus,
        [data-testid="stTextInput"] input:focus {
            border-color: #c9a961 !important;
            box-shadow: 0 0 20px rgba(201, 169, 97, 0.3) !important;
            color: #ffffff !important;
        }

        .stTextInput input::placeholder,
        .stTextInput > div > div > input::placeholder,
        [data-testid="stTextInput"] input::placeholder {
            color: #8892a0 !important;
        }

        /* Selectbox estilizado */
        .stSelectbox > div > div,
        [data-testid="stSelectbox"] > div > div {
            background: rgba(255,255,255,0.1) !important;
            border: 1px solid rgba(201, 169, 97, 0.3) !important;
            border-radius: 8px !important;
            color: #ffffff !important;
        }

        .stSelectbox [data-baseweb="select"] > div,
        [data-testid="stSelectbox"] [data-baseweb="select"] > div {
            background: rgba(255,255,255,0.1) !important;
            color: #ffffff !important;
        }

        .stSelectbox [data-baseweb="select"] span,
        [data-testid="stSelectbox"] [data-baseweb="select"] span {
            color: #ffffff !important;
        }

        /* Checkbox estilizado */
        .stCheckbox label span {
            color: #e6edf3 !important;
        }

        /* ============================================
           SEÇÃO DE CATEGORIAS
           ============================================ */
        .section-title {
            font-size: 22px;
            margin-bottom: 20px;
            display: flex;
            align-items: center;
            gap: 15px;
            color: #fff;
        }

        .section-title::before {
            content: '';
            width: 4px;
            height: 28px;
            background: linear-gradient(180deg, #c9a961 0%, #e6d5a8 100%);
            border-radius: 2px;
        }

        /* ============================================
           FILTROS DIDÁTICOS
           ============================================ */
        .filter-card {
            background: rgba(255,255,255,0.05);
            border: 1px solid rgba(201, 169, 97, 0.2);
            border-radius: 12px;
            padding: 15px;
            margin-bottom: 15px;
        }

        .filter-card-title {
            color: #c9a961;
            font-weight: 600;
            font-size: 14px;
            margin-bottom: 10px;
            display: flex;
            align-items: center;
            gap: 8px;
        }

        .filter-description {
            color: #8892a0;
            font-size: 12px;
            margin-top: 5px;
            font-style: italic;
        }

        /* ============================================
           BADGES DE TRIBUTAÇÃO
           ============================================ */
        .badge-tributacao {
            display: inline-flex;
            align-items: center;
            gap: 6px;
            padding: 6px 12px;
            border-radius: 20px;
            font-size: 12px;
            font-weight: 600;
        }

        .badge-integral {
            background: rgba(76, 175, 80, 0.2);
            color: #4CAF50;
            border: 1px solid #4CAF50;
        }

        .badge-reduzida {
            background: rgba(33, 150, 243, 0.2);
            color: #2196F3;
            border: 1px solid #2196F3;
        }

        .badge-especial {
            background: rgba(255, 152, 0, 0.2);
            color: #FF9800;
            border: 1px solid #FF9800;
        }

        .badge-isento {
            background: rgba(96, 125, 139, 0.2);
            color: #607D8B;
            border: 1px solid #607D8B;
        }

        /* ============================================
           DESTAQUE DE BUSCA
           ============================================ */
        .search-highlight {
            background-color: #FFEB3B !important;
            color: #1a2332 !important;
            padding: 2px 6px;
            border-radius: 4px;
            font-weight: 600;
        }

        /* ============================================
           CÓDIGO MONOESPAÇADO
           ============================================ */
        .code-mono {
            font-family: 'Consolas', 'Monaco', monospace;
            background: rgba(201, 169, 97, 0.15);
            color: #c9a961;
            padding: 3px 8px;
            border-radius: 4px;
            font-size: 13px;
            font-weight: 600;
        }

        /* ============================================
           BOTÕES PREMIUM
           ============================================ */
        .stButton > button {
            background: rgba(255,255,255,0.05) !important;
            border: 1px solid rgba(201, 169, 97, 0.3) !important;
            border-radius: 12px !important;
            padding: 0.8rem 1.2rem !important;
            font-weight: 500 !important;
            transition: all 0.3s !important;
            color: #c9a961 !important;
            min-height: 90px !important;
        }

        .stButton > button:hover {
            background: rgba(201, 169, 97, 0.15) !important;
            border-color: #c9a961 !important;
            transform: translateY(-3px);
            box-shadow: 0 8px 25px rgba(201, 169, 97, 0.25) !important;
        }

        .stButton > button[data-testid="stBaseButton-primary"] {
            background: linear-gradient(135deg, rgba(201, 169, 97, 0.3) 0%, rgba(230, 213, 168, 0.2) 100%) !important;
            border: 2px solid #c9a961 !important;
            box-shadow: 0 0 15px rgba(201, 169, 97, 0.3) !important;
        }

        /* Botão de limpar/exportar */
        .stDownloadButton > button,
        .stButton > button[kind="secondary"] {
            background: transparent !important;
            border: 2px solid #c9a961 !important;
            color: #c9a961 !important;
            min-height: auto !important;
            padding: 0.5rem 1.5rem !important;
        }

        .stDownloadButton > button:hover {
            background: rgba(201, 169, 97, 0.2) !important;
        }

        /* ============================================
           AUTOCOMPLETAR
           ============================================ */
        .autocomplete-item {
            padding: 10px 15px;
            border-bottom: 1px solid rgba(201, 169, 97, 0.1);
            cursor: pointer;
            transition: all 0.2s;
        }

        .autocomplete-item:hover {
            background: rgba(201, 169, 97, 0.15);
        }

        .autocomplete-type {
            font-size: 10px;
            color: #c9a961;
            text-transform: uppercase;
            letter-spacing: 1px;
        }

        /* ============================================
           TABELA DE RESULTADOS
           ============================================ */
        [data-testid="stDataFrame"] {
            background: transparent;
            border-radius: 0;
            padding: 20px;
        }

        [data-testid="stDataFrame"] table {
            border-collapse: separate !important;
            border-spacing: 0 10px !important;
            width: 100% !important;
        }

        /* HEADER */
        [data-testid="stDataFrame"] thead th {
            background: rgba(201, 169, 97, 0.25) !important;
            color: #c9a961 !important;
            font-weight: 700 !important;
            text-transform: uppercase;
            font-size: 13px !important;
            letter-spacing: 1px;
            padding: 16px 15px !important;
            border: none !important;
        }

        [data-testid="stDataFrame"] thead th:first-child {
            border-radius: 8px 0 0 8px !important;
        }

        [data-testid="stDataFrame"] thead th:last-child {
            border-radius: 0 8px 8px 0 !important;
        }

        /* LINHAS */
        [data-testid="stDataFrame"] tbody tr {
            background: rgba(255,255,255,0.05) !important;
            transition: all 0.3s ease;
        }

        [data-testid="stDataFrame"] tbody tr:hover {
            background: rgba(255,255,255,0.12) !important;
            transform: scale(1.005);
            box-shadow: 0 4px 20px rgba(201, 169, 97, 0.2) !important;
        }

        /* CÉLULAS */
        [data-testid="stDataFrame"] tbody td {
            padding: 18px 15px !important;
            border-top: 1px solid rgba(201, 169, 97, 0.15) !important;
            border-bottom: 1px solid rgba(201, 169, 97, 0.15) !important;
            font-size: 14px !important;
            color: #e6edf3 !important;
            vertical-align: middle !important;
            line-height: 1.5 !important;
        }

        [data-testid="stDataFrame"] tbody td:first-child {
            border-left: 1px solid rgba(201, 169, 97, 0.15) !important;
            border-radius: 8px 0 0 8px !important;
            font-weight: 600 !important;
        }

        [data-testid="stDataFrame"] tbody td:last-child {
            border-right: 1px solid rgba(201, 169, 97, 0.15) !important;
            border-radius: 0 8px 8px 0 !important;
        }

        /* ============================================
           INFO BOXES
           ============================================ */
        .info-box {
            background: rgba(201, 169, 97, 0.1);
            border: 1px solid rgba(201, 169, 97, 0.3);
            border-radius: 10px;
            padding: 15px 20px;
            margin: 15px 0;
        }

        .info-box-title {
            color: #c9a961;
            font-weight: 700;
            font-size: 14px;
            margin-bottom: 8px;
        }

        .info-box-content {
            color: #e6edf3;
            font-size: 13px;
            line-height: 1.6;
        }

        /* ============================================
           LEGENDA DE CORES
           ============================================ */
        .legenda-item {
            display: inline-flex;
            align-items: center;
            gap: 6px;
            margin-right: 15px;
            font-size: 12px;
        }

        .legenda-cor {
            width: 12px;
            height: 12px;
            border-radius: 3px;
        }

        /* ============================================
           MENSAGEM SEM RESULTADOS
           ============================================ */
        .no-results {
            text-align: center;
            padding: 60px 20px;
            color: #8892a0;
        }

        .no-results-icon {
            font-size: 48px;
            margin-bottom: 20px;
        }

        .no-results h3 {
            color: #c9a961;
            margin-bottom: 10px;
        }

        /* ============================================
           RESULTADOS INFO
           ============================================ */
        .results-info {
            color: #b0b8c1;
            font-size: 14px;
            margin-top: 5px;
        }

        /* ============================================
           ESCONDER ELEMENTOS
           ============================================ */
        #MainMenu {visibility: hidden;}
        footer {visibility: hidden;}
    </style>
    """, unsafe_allow_html=True)


# =============================================================================
# MAPEAMENTO DE ÍCONES DAS CATEGORIAS
# =============================================================================

CATEGORY_ICONS = {
    "1. SERVIÇOS FINANCEIROS E BANCÁRIOS": "💼",
    "2. SERVIÇOS DE SAÚDE HUMANA": "🏥",
    "3. SERVIÇOS JURÍDICOS E CONTÁBEIS": "⚖️",
    "4. ENGENHARIA E CONSTRUÇÃO CIVIL": "🏗️",
    "5. TECNOLOGIA DA INFORMAÇÃO": "💻",
    "6. CONSULTORIA E ASSESSORIA": "💡",
    "7. TRANSPORTE E LOGÍSTICA": "🚚",
    "8. EDUCAÇÃO E TREINAMENTO": "📚",
    "9. COMUNICAÇÃO, PUBLICIDADE E EVENTOS": "📢",
    "10. ENTRETENIMENTO E LAZER": "🎭",
    "11. BELEZA E ESTÉTICA": "✨",
    "12. LIMPEZA E CONSERVAÇÃO": "🧹",
    "13. SEGURANÇA": "🔒",
    "14. MANUTENÇÃO E REPARAÇÃO": "🔧",
    "15. RECURSOS HUMANOS": "👥",
    "16. OUTROS SERVIÇOS": "🎯",
}


# =============================================================================
# FUNÇÕES DE EXPORTAÇÃO
# =============================================================================

def export_to_excel(results, search_term=None):
    """Exporta resultados para Excel com formatação profissional."""
    export_data = []
    for item in results:
        for nbs in item.get('nbs_entries', []):
            classificacoes = nbs.get('cclasstrib', [])

            if classificacoes:
                class_info = classificacoes[0]
                cod_class = class_info.get('codigo', '')
                nome_class = class_info.get('nome', '')
                class_completa = f"{cod_class} - {nome_class}"
            else:
                class_completa = "-"

            export_data.append({
                'Código LC116': item.get('item_lc116', ''),
                'Descrição do Serviço': item.get('descricao_item', ''),
                'Código NBS': nbs.get('nbs_code', ''),
                'Descrição NBS': nbs.get('descricao_nbs', ''),
                'Prestação Onerosa': 'Sim' if nbs.get('ps_onerosa') == 'S' else 'Não' if nbs.get('ps_onerosa') == 'N' else '-',
                'Aquisição Exterior': 'Sim' if nbs.get('adq_exterior') == 'S' else 'Não' if nbs.get('adq_exterior') == 'N' else '-',
                'INDOP': nbs.get('indop', '-'),
                'Local Incidência IBS': nbs.get('local_incidencia_ibs', '-'),
                'Classificação Tributária': class_completa,
            })

    if not export_data:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Consulta Tributária"

    headers = list(export_data[0].keys())
    ws.append(headers)

    for row_data in export_data:
        ws.append(list(row_data.values()))

    header_fill = PatternFill(start_color="C9A961", end_color="C9A961", fill_type="solid")
    header_font = Font(name='Calibri', size=11, bold=True, color="1A2332")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    data_font = Font(name='Calibri', size=10)
    data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin', color='C9A961'),
        right=Side(style='thin', color='C9A961'),
        top=Side(style='thin', color='C9A961'),
        bottom=Side(style='thin', color='C9A961')
    )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

    column_widths = {'A': 15, 'B': 50, 'C': 18, 'D': 50, 'E': 18, 'F': 18, 'G': 15, 'H': 40, 'I': 60}
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    tab = Table(displayName="TabelaTributaria", ref=f"A1:{get_column_letter(len(headers))}{len(export_data)+1}")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    ws.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()


# =============================================================================
# COMPONENTES DE UI
# =============================================================================

def render_header():
    """Renderiza o cabeçalho premium."""
    st.markdown("""
    <div class="premium-header">
        <div class="logo-circle">NC</div>
        <div>
            <div class="header-title">Sistema de Consulta Tributária</div>
            <div class="header-subtitle">IBS/CBS - Classificações Tributárias | Neto Contabilidade</div>
        </div>
    </div>
    """, unsafe_allow_html=True)


def render_search_hero(search_service, items):
    """Renderiza a seção hero de busca com autocompletar."""
    st.markdown("""
    <div class="search-hero">
        <div class="search-title">Encontre o Código Tributário Correto</div>
        <div class="search-subtitle">Digite o código LC116, NBS, descrição do serviço ou palavras-chave relacionadas</div>
    </div>
    """, unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 3, 1])
    with col2:
        search_term = st.text_input(
            "Busca",
            placeholder="Ex: desenvolvimento de software, 1.01, consultoria, TI...",
            label_visibility="collapsed",
            key="main_search"
        )

        # Exibir sugestões de autocompletar
        if search_term and len(search_term) >= 2:
            suggestions = search_service.get_autocomplete_suggestions(items, search_term, max_suggestions=6)
            if suggestions:
                with st.expander(f"💡 {len(suggestions)} sugestões encontradas", expanded=True):
                    for sug in suggestions:
                        tipo_badge = f"<span style='font-size:10px;color:#c9a961;'>[{sug['tipo']}]</span>"
                        st.markdown(f"{tipo_badge} {sug['texto']}", unsafe_allow_html=True)

        # Opções de busca
        col_opt1, col_opt2, col_opt3 = st.columns([1, 1, 1])
        with col_opt1:
            search_type = st.selectbox(
                "Tipo de Busca",
                ["Contém", "Aproximada (Fuzzy)", "Exata"],
                label_visibility="collapsed",
                key="search_type",
                help="Contém: busca parcial | Aproximada: tolera erros de digitação | Exata: match preciso"
            )
        with col_opt2:
            use_synonyms = st.checkbox("🔗 Usar sinônimos", value=True, key="use_synonyms",
                                       help="Expande a busca incluindo termos relacionados")
        with col_opt3:
            sort_option = st.selectbox(
                "Ordenar por",
                ["Relevância", "Código LC116", "Código NBS"],
                label_visibility="collapsed",
                key="sort_option"
            )

    type_map = {"Contém": "contains", "Aproximada (Fuzzy)": "fuzzy", "Exata": "exact"}
    return search_term, type_map.get(search_type, "contains"), use_synonyms, sort_option


def render_category_grid(items, search_service):
    """Renderiza o grid de categorias clicáveis."""
    if 'selected_categoria' not in st.session_state:
        st.session_state.selected_categoria = None
    if 'selected_subcategoria' not in st.session_state:
        st.session_state.selected_subcategoria = None

    categoria_counts = {}
    for item in items:
        cat = item.get('filtro_principal', '')
        if cat:
            nbs_count = len(item.get('nbs_entries', []))
            categoria_counts[cat] = categoria_counts.get(cat, 0) + nbs_count

    st.markdown('<div class="section-title">Categorias Principais</div>', unsafe_allow_html=True)

    if st.session_state.selected_categoria:
        col_clear = st.columns([1, 2, 1])[1]
        with col_clear:
            if st.button("🔄 Limpar Seleção de Categoria", use_container_width=True, key="clear_cat"):
                st.session_state.selected_categoria = None
                st.session_state.selected_subcategoria = None
                st.rerun()

    categorias = sorted(CATEGORY_ICONS.keys())
    outros_key = "16. OUTROS SERVIÇOS"
    if outros_key in categorias:
        categorias.remove(outros_key)
        categorias.append(outros_key)

    num_cols = 4
    rows = [categorias[i:i+num_cols] for i in range(0, len(categorias), num_cols)]

    for row in rows:
        cols = st.columns(num_cols)
        for idx, cat in enumerate(row):
            with cols[idx]:
                count = categoria_counts.get(cat, 0)
                icon = CATEGORY_ICONS.get(cat, "📋")
                is_selected = st.session_state.selected_categoria == cat
                cat_name = cat.split(". ", 1)[1] if ". " in cat else cat
                btn_label = f"{icon} {cat_name}\n({count} itens)"

                if st.button(btn_label, key=f"cat_{cat}", use_container_width=True,
                            type="primary" if is_selected else "secondary"):
                    if st.session_state.selected_categoria == cat:
                        st.session_state.selected_categoria = None
                        st.session_state.selected_subcategoria = None
                    else:
                        st.session_state.selected_categoria = cat
                        st.session_state.selected_subcategoria = None
                    st.rerun()

    selected_subcategoria = None
    if st.session_state.selected_categoria:
        st.markdown("---")
        cat_display = st.session_state.selected_categoria.split(". ", 1)[1] if ". " in st.session_state.selected_categoria else st.session_state.selected_categoria
        st.markdown(f'<div class="section-title">📂 Subcategorias de {cat_display}</div>', unsafe_allow_html=True)

        subcategorias = search_service.get_subcategorias_by_filtro(items, st.session_state.selected_categoria)

        if subcategorias:
            sub_counts = {}
            for item in items:
                if item.get('filtro_principal') == st.session_state.selected_categoria:
                    sub = item.get('subcategoria', '')
                    if sub:
                        nbs_count = len(item.get('nbs_entries', []))
                        sub_counts[sub] = sub_counts.get(sub, 0) + nbs_count

            col_all = st.columns([1, 2, 1])[1]
            with col_all:
                total = sum(sub_counts.values())
                if st.button(f"📋 Todas as Subcategorias ({total})", key="sub_all", use_container_width=True,
                            type="primary" if st.session_state.selected_subcategoria is None else "secondary"):
                    st.session_state.selected_subcategoria = None
                    st.rerun()

            num_sub_cols = 3
            sub_rows = [subcategorias[i:i+num_sub_cols] for i in range(0, len(subcategorias), num_sub_cols)]

            for sub_row in sub_rows:
                sub_cols = st.columns(num_sub_cols)
                for idx, sub in enumerate(sub_row):
                    with sub_cols[idx]:
                        count = sub_counts.get(sub, 0)
                        is_sel = st.session_state.selected_subcategoria == sub
                        if st.button(f"📂 {sub} ({count})", key=f"sub_{sub}", use_container_width=True,
                                    type="primary" if is_sel else "secondary"):
                            if st.session_state.selected_subcategoria == sub:
                                st.session_state.selected_subcategoria = None
                            else:
                                st.session_state.selected_subcategoria = sub
                            st.rerun()

            selected_subcategoria = st.session_state.selected_subcategoria

    return st.session_state.selected_categoria, selected_subcategoria


def render_tributacao_badge(codigo: str, search_service: SearchServiceEnhanced) -> str:
    """Retorna HTML para badge de tributação."""
    info = search_service.get_classificacao_didatica(codigo)
    return f"""
    <span class="badge-tributacao" style="background: {info['cor']}22; color: {info['cor']}; border: 1px solid {info['cor']};">
        {info['icone']} {info['categoria']}
    </span>
    """


def render_detailed_view(results, search_service, search_term=None):
    """Renderiza visualização detalhada com cards expandíveis e destaque de busca."""
    st.markdown("""
    <div class='info-box'>
        <div class='info-box-title'>📋 Visualização Detalhada</div>
        <div class='info-box-content'>
            Clique em cada serviço para expandir e ver todos os códigos NBS relacionados, 
            classificações tributárias e informações detalhadas.
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Legenda de cores
    st.markdown("""
    <div style="margin-bottom: 20px;">
        <strong style="color: #c9a961;">Legenda de Tributação:</strong><br>
        <span class="legenda-item"><span class="legenda-cor" style="background: #4CAF50;"></span> Tributação Integral</span>
        <span class="legenda-item"><span class="legenda-cor" style="background: #2196F3;"></span> Alíquota Reduzida</span>
        <span class="legenda-item"><span class="legenda-cor" style="background: #FF9800;"></span> Regime Especial</span>
        <span class="legenda-item"><span class="legenda-cor" style="background: #607D8B;"></span> Isenção</span>
    </div>
    """, unsafe_allow_html=True)

    for item in results:
        lc116 = item.get('item_lc116', '')
        desc_servico = item.get('descricao_item', '')
        nbs_entries = item.get('nbs_entries', [])

        # Aplicar destaque de busca se houver termo
        desc_display = desc_servico
        if search_term:
            desc_display = search_service.highlight_text(desc_servico, search_term, "#FFEB3B")

        with st.expander(f"**{lc116}** - {desc_servico[:80]}...", expanded=False):
            st.markdown(f"""
            <div style='background: rgba(45, 62, 80, 0.5); padding: 15px; border-radius: 8px; margin-bottom: 15px;'>
                <div style='color: #c9a961; font-weight: 700; font-size: 16px; margin-bottom: 8px;'>
                    <span class='code-mono'>{lc116}</span> Serviço LC116
                </div>
                <div style='color: #e6edf3; font-size: 14px; line-height: 1.6;'>
                    {desc_display}
                </div>
                <div style='margin-top: 10px; color: #b0b8c1; font-size: 13px;'>
                    <strong>{len(nbs_entries)}</strong> códigos NBS vinculados
                </div>
            </div>
            """, unsafe_allow_html=True)

            for idx, nbs in enumerate(nbs_entries, 1):
                classificacoes = nbs.get('cclasstrib', [])
                nbs_desc = nbs.get('descricao_nbs', '')
                
                # Destaque na descrição NBS
                if search_term:
                    nbs_desc = search_service.highlight_text(nbs_desc, search_term, "#FFEB3B")

                # Badges de classificação
                badges_html = ""
                for class_info in classificacoes:
                    codigo = class_info.get('codigo', '')
                    badges_html += render_tributacao_badge(codigo, search_service)

                # Formatação de Prestação Onerosa e Aquisição Exterior
                ps_onerosa = nbs.get('ps_onerosa', '')
                onerosa_display = '✅ Sim' if ps_onerosa == 'S' else '❌ Não' if ps_onerosa == 'N' else '➖'
                
                adq_ext = nbs.get('adq_exterior', '')
                exterior_display = '✅ Sim' if adq_ext == 'S' else '❌ Não' if adq_ext == 'N' else '➖'

                st.markdown(f"""
                <div style='background: linear-gradient(135deg, rgba(201, 169, 97, 0.2) 0%, rgba(230, 213, 168, 0.1) 100%);
                            border-left: 4px solid #c9a961; padding: 15px; border-radius: 8px; margin-bottom: 12px;'>
                    <div style='display: flex; justify-content: space-between; align-items: center; margin-bottom: 10px;'>
                        <div>
                            <span style='color: #c9a961; font-weight: 700; font-size: 15px;'>
                                NBS: <span class='code-mono'>{nbs.get('nbs_code', '')}</span>
                            </span>
                            <span style='margin-left: 10px;'>{badges_html}</span>
                        </div>
                        <div style='color: #8892a0; font-size: 12px;'>
                            #{idx} de {len(nbs_entries)}
                        </div>
                    </div>
                    <div style='color: #e6edf3; font-size: 13px; margin-bottom: 12px; line-height: 1.5;'>
                        {nbs_desc}
                    </div>
                    <div style='display: grid; grid-template-columns: repeat(2, 1fr); gap: 10px;'>
                        <div style='background: rgba(26, 35, 50, 0.5); padding: 8px; border-radius: 6px;'>
                            <div style='color: #8892a0; font-size: 11px; margin-bottom: 4px;'>💰 Prestação Onerosa</div>
                            <div style='color: #e6d5a8; font-weight: 600; font-size: 13px;'>{onerosa_display}</div>
                        </div>
                        <div style='background: rgba(26, 35, 50, 0.5); padding: 8px; border-radius: 6px;'>
                            <div style='color: #8892a0; font-size: 11px; margin-bottom: 4px;'>🌍 Aquisição Exterior</div>
                            <div style='color: #e6d5a8; font-weight: 600; font-size: 13px;'>{exterior_display}</div>
                        </div>
                        <div style='background: rgba(26, 35, 50, 0.5); padding: 8px; border-radius: 6px;'>
                            <div style='color: #8892a0; font-size: 11px; margin-bottom: 4px;'>🔢 INDOP</div>
                            <div style='color: #e6d5a8; font-weight: 600; font-size: 13px; font-family: monospace;'>
                                {nbs.get('indop', '-')}
                            </div>
                        </div>
                        <div style='background: rgba(26, 35, 50, 0.5); padding: 8px; border-radius: 6px;'>
                            <div style='color: #8892a0; font-size: 11px; margin-bottom: 4px;'>📍 Local Incidência IBS</div>
                            <div style='color: #e6d5a8; font-weight: 600; font-size: 12px;'>
                                {nbs.get('local_incidencia_ibs', '-')}
                            </div>
                        </div>
                    </div>
                """, unsafe_allow_html=True)

                if classificacoes:
                    st.markdown("""
                    <div style='margin-top: 10px; padding: 10px; background: rgba(201, 169, 97, 0.15); border-radius: 6px;'>
                        <div style='color: #c9a961; font-weight: 700; font-size: 13px; margin-bottom: 8px;'>
                            🏛️ Classificações Tributárias:
                        </div>
                    """, unsafe_allow_html=True)

                    for class_info in classificacoes:
                        codigo = class_info.get('codigo', '')
                        nome = class_info.get('nome', '')
                        info_didatica = search_service.get_classificacao_didatica(codigo)
                        
                        st.markdown(f"""
                        <div style='margin-left: 10px; margin-bottom: 8px; padding: 8px; background: rgba(26, 35, 50, 0.3); border-radius: 4px;'>
                            <div style='display: flex; align-items: center; gap: 8px;'>
                                <span style='font-size: 16px;'>{info_didatica['icone']}</span>
                                <span class='code-mono'>{codigo}</span>
                                <span style='color: {info_didatica["cor"]}; font-weight: 600; font-size: 12px;'>
                                    {info_didatica['categoria']}
                                </span>
                            </div>
                            <div style='color: #e6edf3; font-size: 12px; margin-top: 5px;'>
                                {nome}
                            </div>
                        </div>
                        """, unsafe_allow_html=True)

                    st.markdown("</div>", unsafe_allow_html=True)

                st.markdown("</div>", unsafe_allow_html=True)


def render_results_table(results, data_service, search_service, search_term=None, sort_option="Relevância"):
    """Renderiza a tabela de resultados com destaque de busca e ordenação."""
    if not results:
        st.markdown("""
        <div class="no-results">
            <div class="no-results-icon">🔍</div>
            <h3>Nenhum resultado encontrado</h3>
            <p>Tente ajustar os filtros ou termos de busca. Dica: use sinônimos ou termos mais gerais.</p>
        </div>
        """, unsafe_allow_html=True)
        return

    total_nbs = sum(len(item.get('nbs_entries', [])) for item in results)

    col1, col2 = st.columns([3, 1])
    with col1:
        st.markdown(f'<div class="section-title">Resultados da Busca</div>', unsafe_allow_html=True)
        st.markdown(f'<div class="results-info">{len(results)} serviços, {total_nbs} entradas NBS encontradas</div>', unsafe_allow_html=True)

    with col2:
        excel_data = export_to_excel(results, search_term)
        if excel_data:
            filename = f"consulta_tributaria_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            st.download_button(
                label="📊 Exportar Excel",
                data=excel_data,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

    st.markdown("<br>", unsafe_allow_html=True)

    # Ordenação
    if sort_option == "Código LC116":
        results = sorted(results, key=lambda x: x.get('item_lc116', ''))
    elif sort_option == "Código NBS":
        results = sorted(results, key=lambda x: x.get('nbs_entries', [{}])[0].get('nbs_code', '') if x.get('nbs_entries') else '')

    # Preparar dados para DataFrame
    table_data = []
    for item in results:
        for nbs in item.get('nbs_entries', []):
            classificacoes = nbs.get('cclasstrib', [])

            if classificacoes:
                class_info = classificacoes[0]
                class_display = class_info.get('codigo', '')
                class_nome = class_info.get('nome', '')
                info_didatica = search_service.get_classificacao_didatica(class_display)
                tipo_trib = f"{info_didatica['icone']} {info_didatica['categoria']}"
            else:
                class_display = "-"
                class_nome = "-"
                tipo_trib = "-"

            desc_servico = item.get('descricao_item', '')
            if len(desc_servico) > 50:
                desc_servico = desc_servico[:50] + '...'

            desc_nbs_text = nbs.get('descricao_nbs', '')
            if len(desc_nbs_text) > 50:
                desc_nbs_text = desc_nbs_text[:50] + '...'

            prest_onerosa = nbs.get('ps_onerosa', '')
            prest_onerosa_display = '✅' if prest_onerosa == 'S' else '❌' if prest_onerosa == 'N' else '➖'

            aquis_ext = nbs.get('adq_exterior', '')
            aquis_ext_display = '✅' if aquis_ext == 'S' else '❌' if aquis_ext == 'N' else '➖'

            local_incid = nbs.get('local_incidencia_ibs', '')
            if len(local_incid) > 35:
                local_incid = local_incid[:35] + '...'

            table_data.append({
                'LC116': item.get("item_lc116", ""),
                'Serviço': desc_servico,
                'NBS': nbs.get("nbs_code", ""),
                'Desc. NBS': desc_nbs_text,
                'Onerosa': prest_onerosa_display,
                'Exterior': aquis_ext_display,
                'cClassTrib': class_display,
                'Tipo Trib.': tipo_trib,
                'Local IBS': local_incid,
            })

    if table_data:
        df = pd.DataFrame(table_data)

        # Tabs para visualização
        tab1, tab2 = st.tabs(["📊 Tabela Completa", "📋 Visualização Detalhada"])

        with tab1:
            st.dataframe(
                df,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "LC116": st.column_config.TextColumn("LC116", width="small"),
                    "Serviço": st.column_config.TextColumn("Serviço", width="medium"),
                    "NBS": st.column_config.TextColumn("NBS", width="medium"),
                    "Desc. NBS": st.column_config.TextColumn("Desc. NBS", width="medium"),
                    "Onerosa": st.column_config.TextColumn("Onerosa", width="small"),
                    "Exterior": st.column_config.TextColumn("Exterior", width="small"),
                    "cClassTrib": st.column_config.TextColumn("cClassTrib", width="small"),
                    "Tipo Trib.": st.column_config.TextColumn("Tipo Trib.", width="medium"),
                    "Local IBS": st.column_config.TextColumn("Local IBS", width="medium"),
                },
                height=600
            )

        with tab2:
            render_detailed_view(results, search_service, search_term)


def render_sidebar_filters(data_service, search_service, items, selected_categoria=None):
    """Renderiza filtros avançados na sidebar com descrições didáticas."""
    filters = data_service.filters

    st.sidebar.markdown("## ⚙️ Filtros Avançados")

    # Botão limpar
    if st.sidebar.button("🔄 Limpar Todos os Filtros", use_container_width=True):
        for key in list(st.session_state.keys()):
            if key.startswith('filtro_') or key.startswith('selected_'):
                del st.session_state[key]
        st.rerun()

    st.sidebar.markdown("---")

    # ========================================
    # FILTRO DIDÁTICO: Tipo de Tributação
    # ========================================
    st.sidebar.markdown("""
    <div class="filter-card">
        <div class="filter-card-title">🏛️ Tipo de Tributação</div>
        <div class="filter-description">
            Filtre por categoria tributária para encontrar serviços com tratamento específico.
        </div>
    </div>
    """, unsafe_allow_html=True)

    tipos_trib = search_service.get_tipos_tributacao_disponiveis(items)
    opcoes_trib = ["Todos"] + [t['nome'] for t in tipos_trib]
    
    tipo_trib_selected = st.sidebar.selectbox(
        "Tipo de Tributação",
        opcoes_trib,
        key="filtro_tipo_tributacao",
        label_visibility="collapsed"
    )
    selected_tipo_trib = None if tipo_trib_selected == "Todos" else tipo_trib_selected

    # Mostrar descrição do tipo selecionado
    if selected_tipo_trib:
        for t in tipos_trib:
            if t['nome'] == selected_tipo_trib:
                st.sidebar.info(f"{t['icone']} {t['descricao']}")
                break

    st.sidebar.markdown("---")

    # ========================================
    # FILTRO: Grupo LC116
    # ========================================
    st.sidebar.markdown("""
    <div class="filter-card">
        <div class="filter-card-title">📂 Grupo de Serviços (LC 116)</div>
        <div class="filter-description">
            Navegue por grupos temáticos da Lei Complementar 116/2003.
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Se uma categoria está selecionada, desabilitar este filtro para evitar conflitos
    if selected_categoria:
        st.sidebar.info("⚠️ Filtro de grupo desabilitado quando uma categoria está selecionada.")
        selected_grupo = None
    else:
        grupos_lc116 = search_service.get_grupos_lc116_disponiveis(items)
        opcoes_grupos = ["Todos"] + [g['display'] for g in grupos_lc116]
        
        grupo_selected = st.sidebar.selectbox(
            "Grupo LC116",
            opcoes_grupos,
            key="filtro_grupo_lc116",
            label_visibility="collapsed"
        )
        selected_grupo = None if grupo_selected == "Todos" else grupo_selected.split(' - ')[0]

    st.sidebar.markdown("---")

    # ========================================
    # FILTRO: Prestação Onerosa
    # ========================================
    with st.sidebar.expander("💰 Prestação Onerosa", expanded=False):
        st.markdown("""
        <div class="filter-description">
            Indica se o serviço é prestado mediante pagamento.
        </div>
        """, unsafe_allow_html=True)
        onerosa = st.radio(
            "Selecione",
            ["Todas", "S", "N"],
            format_func=lambda x: {"Todas": "📋 Todas", "S": "✅ Sim (Onerosa)", "N": "❌ Não (Gratuita)"}.get(x, x),
            key="filtro_onerosa",
            horizontal=True,
            label_visibility="collapsed"
        )
    selected_onerosa = None if onerosa == "Todas" else onerosa

    # ========================================
    # FILTRO: Aquisição Exterior
    # ========================================
    with st.sidebar.expander("🌍 Aquisição Exterior", expanded=False):
        st.markdown("""
        <div class="filter-description">
            Indica se o serviço pode envolver importação (aquisição do exterior).
        </div>
        """, unsafe_allow_html=True)
        exterior = st.radio(
            "Selecione",
            ["Todas", "S", "N"],
            format_func=lambda x: {"Todas": "📋 Todas", "S": "✅ Sim (Importação)", "N": "❌ Não"}.get(x, x),
            key="filtro_exterior",
            horizontal=True,
            label_visibility="collapsed"
        )
    selected_exterior = None if exterior == "Todas" else exterior

    # ========================================
    # FILTRO: Local de Incidência
    # ========================================
    with st.sidebar.expander("📍 Local de Incidência do IBS", expanded=False):
        st.markdown("""
        <div class="filter-description">
            Define onde o imposto é devido: no destino (local do tomador) ou origem (local do prestador).
        </div>
        """, unsafe_allow_html=True)
        locais = ["Todos"] + filters.get('local_incidencia', [])
        local = st.selectbox("Selecione", locais, key="filtro_local", label_visibility="collapsed")
    selected_local = None if local == "Todos" else local

    # ========================================
    # FILTRO: Classificação Tributária Específica
    # ========================================
    with st.sidebar.expander("🔢 Classificação Tributária (cClassTrib)", expanded=False):
        st.markdown("""
        <div class="filter-description">
            Filtre por código específico de classificação tributária.
        </div>
        """, unsafe_allow_html=True)
        classificacoes = ["Todas"] + filters.get('classificacoes_tributarias', [])
        classificacao = st.selectbox("Selecione", classificacoes, key="filtro_classificacao", label_visibility="collapsed")
    selected_classificacao = None if classificacao == "Todas" else classificacao

    st.sidebar.markdown("---")

    # Links úteis
    st.sidebar.markdown("### 📚 Legislação e Documentação")
    st.sidebar.markdown("""
    - [LC 116/2003](https://www.planalto.gov.br/ccivil_03/leis/lcp/lcp116.htm)
    - [Receita Federal](https://www.gov.br/receitafederal)
    - [Reforma Tributária](https://www.gov.br/fazenda/pt-br/assuntos/reforma-tributaria)
    """)

    st.sidebar.markdown("---")
    st.sidebar.markdown("""
    <div style="text-align: center; color: #8892a0; font-size: 11px;">
        <strong>Neto Contabilidade</strong><br>
        Fonte: AnexoVIII Correlação v1.00<br>
        Sistema v2.0.0 | 2025
    </div>
    """, unsafe_allow_html=True)

    return {
        'ps_onerosa': selected_onerosa,
        'adq_exterior': selected_exterior,
        'local_incidencia': selected_local,
        'cclasstrib_filter': selected_classificacao,
        'tipo_tributacao': selected_tipo_trib,
        'grupo_lc116': selected_grupo,
    }


# =============================================================================
# FUNÇÃO PRINCIPAL
# =============================================================================

def main():
    """Função principal da aplicação."""
    configure_page()
    render_header()

    # Inicializa serviços
    data_service = DataService(DATA_FILE)
    search_service = SearchServiceEnhanced(fuzzy_threshold=SEARCH_CONFIG["fuzzy_threshold"])

    # Carrega dados
    if not data_service.load_data():
        st.error("❌ Falha ao carregar os dados. Verifique se o arquivo JSON está disponível.")
        st.stop()

    items = data_service.items

    # Inicializar estado da categoria (para passar aos filtros da sidebar)
    if 'selected_categoria' not in st.session_state:
        st.session_state.selected_categoria = None

    # Filtros na sidebar (passando a categoria selecionada para desabilitar filtro de grupo se necessário)
    sidebar_filters = render_sidebar_filters(data_service, search_service, items, st.session_state.selected_categoria)

    # Hero de busca com autocompletar
    search_term, search_type, use_synonyms, sort_option = render_search_hero(search_service, items)

    # Grid de categorias
    selected_categoria, selected_subcategoria = render_category_grid(items, search_service)

    st.markdown("---")

    # Aplicar busca
    results = items

    if search_term and len(search_term) >= SEARCH_CONFIG["min_search_length"]:
        results = search_service.search_items(
            results,
            search_term,
            search_type=search_type,
            use_synonyms=use_synonyms
        )

    # Aplicar filtros
    results = search_service.filter_items(
        results,
        filtro_principal=selected_categoria,
        subcategoria=selected_subcategoria,
        ps_onerosa=sidebar_filters.get('ps_onerosa'),
        adq_exterior=sidebar_filters.get('adq_exterior'),
        local_incidencia=sidebar_filters.get('local_incidencia'),
        cclasstrib_filter=sidebar_filters.get('cclasstrib_filter'),
        tipo_tributacao=sidebar_filters.get('tipo_tributacao'),
        grupo_lc116=sidebar_filters.get('grupo_lc116'),
    )

    # Tabela de resultados
    render_results_table(results, data_service, search_service, search_term, sort_option)


if __name__ == "__main__":
    main()
